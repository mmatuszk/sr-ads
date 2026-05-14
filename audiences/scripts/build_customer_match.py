#!/usr/bin/env python3
"""Build a Google Customer Match review workbook from Shopify orders.

The source file is expected to be a Matrixify/Shopify order export with an
`Orders` sheet and standard order, customer, and line-item columns.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl
import yaml
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_CONFIG = Path("audiences/config/customer_match/workflow.yaml")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
TOP_VALUES = {True, "TRUE", "true", "Y", "Yes", 1, "1"}
RULE_KEYS = (
    "lookback_years",
    "min_recent_spend",
    "repeat_order_min",
    "large_single_order",
    "lifetime_spend",
    "lifetime_recent_spend",
    "max_customers",
)
DEFAULT_RULES = {
    "lookback_years": 3,
    "min_recent_spend": 250,
    "repeat_order_min": 2,
    "large_single_order": 500,
    "lifetime_spend": 3000,
    "lifetime_recent_spend": 100,
    "max_customers": None,
}
DEFAULT_COLUMNS = {
    "top_row": "Top Row",
    "processed_at": "Processed At",
    "email": "Email",
    "customer_email": "Customer: Email",
    "current_total": "Price: Current Total",
    "total": "Price: Total",
    "payment_status": "Payment: Status",
    "cancelled_at": "Cancelled At",
    "order_id": "ID",
    "customer_tags": "Customer: Tags",
    "order_tags": "Tags",
    "customer_total_spent": "Customer: Total Spent",
    "customer_orders_count": "Customer: Orders Count",
    "line_type": "Line: Type",
    "line_title": "Line: Title",
    "line_name": "Line: Name",
    "line_sku": "Line: SKU",
    "line_product_handle": "Line: Product Handle",
    "line_vendor": "Line: Vendor",
    "line_total": "Line: Total",
}


class ConfigError(ValueError):
    pass


def parse_dt(value: Any) -> dt.datetime | None:
    if not value:
        return None
    if isinstance(value, dt.datetime):
        return value
    return dt.datetime.strptime(str(value)[:19], "%Y-%m-%d %H:%M:%S")


def money(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def norm_email(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def split_tags(value: Any) -> set[str]:
    if not value:
        return set()
    return {part.strip().lower() for part in re.split(r"[,;]", str(value)) if part.strip()}


def row_value(row: tuple[Any, ...], idx: dict[str, int], columns: dict[str, str], key: str) -> Any:
    return row[idx[columns[key]]]


def classify_line(row: tuple[Any, ...], idx: dict[str, int], columns: dict[str, str]) -> set[str]:
    text = " ".join(
        str(row_value(row, idx, columns, key) or "").lower()
        for key in ("line_title", "line_name", "line_sku", "line_product_handle", "line_vendor")
    )
    line_total = money(row_value(row, idx, columns, "line_total"))
    classes = set()

    if "sample" in text or (0 < line_total < 30):
        classes.add("sample")
    if "wallpaper" in text or "double roll" in text or "single roll" in text:
        classes.add("wallpaper")
    if any(term in text for term in ("fabric", "silk", "linen", "cotton", "velvet", "damask", "jacquard", "toile")):
        classes.add("fabric")
    if any(term in text for term in ("trim", "tape", "cord", "fringe", "samuel & sons")):
        classes.add("trim")
    if any(term in text for term in ("pillow", "bolster")):
        classes.add("pillow")
    if "scalamandre" in text or "old world weavers" in text:
        classes.add("scalamandre")

    return classes


def build_reason(customer: dict[str, Any], rules: dict[str, Any]) -> str:
    reasons = []

    if "designer-program" in customer["tags"]:
        reasons.append("designer-program customer")
    if customer["spend"] >= 1000:
        reasons.append(f"high recent spend ${customer['spend']:,.0f}")
    elif customer["spend"] >= rules["min_recent_spend"]:
        reasons.append(f"qualified recent spend ${customer['spend']:,.0f}")
    elif customer["orders"] >= rules["repeat_order_min"]:
        reasons.append(f"repeat buyer with {customer['orders']} recent orders")
    elif customer["max_order"] >= rules["large_single_order"]:
        reasons.append(f"large single order ${customer['max_order']:,.0f}")

    if customer["lifetime_spent"] >= rules["lifetime_spend"] and customer["spend"] >= rules["lifetime_recent_spend"]:
        reasons.append(f"strong lifetime value ${customer['lifetime_spent']:,.0f}")
    elif customer["lifetime_orders"] >= 3 and customer["spend"] >= rules["lifetime_recent_spend"]:
        reasons.append(f"{customer['lifetime_orders']} lifetime orders")

    classes = customer["line_classes"]
    if classes["fabric"]:
        reasons.append("fabric buyer")
    if classes["wallpaper"]:
        reasons.append("wallpaper buyer")
    if classes["scalamandre"]:
        reasons.append("Scalamandre/design-house interest")
    if classes["trim"]:
        reasons.append("trim buyer")
    if classes["pillow"]:
        reasons.append("pillow buyer")

    reasons.append(f"last order {customer['last'].date().isoformat()}")
    return "; ".join(reasons)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG, help="Audience workflow config YAML")
    parser.add_argument("--preset", default="baseline", help="Named preset from the workflow config")
    parser.add_argument("--run", type=Path, help="Dated run directory containing source/ and output/")
    parser.add_argument("--input", type=Path, help="Path to Matrixify/Shopify order export .xlsx")
    parser.add_argument("--output", type=Path, help="Path to write the two-column .xlsx output")
    parser.add_argument("--csv-output", type=Path, help="Optional two-column CSV output path")
    parser.add_argument("--google-upload-output", type=Path, help="Optional Google Ads upload-ready CSV output path")
    parser.add_argument("--pinterest-upload-output", type=Path, help="Optional Pinterest Ads upload-ready CSV output path")
    parser.add_argument("--summary-output", type=Path, help="Optional JSON summary output path")
    parser.add_argument("--lookback-years", type=int)
    parser.add_argument("--min-recent-spend", type=float)
    parser.add_argument("--repeat-order-min", type=int)
    parser.add_argument("--large-single-order", type=float)
    parser.add_argument("--lifetime-spend", type=float)
    parser.add_argument("--lifetime-recent-spend", type=float)
    parser.add_argument(
        "--max-customers",
        type=int,
        help="Optional cap after sorting selected customers by spend, order count, and email",
    )
    parser.add_argument("--exclude-email", action="append", default=[], help="Email address to exclude; repeatable")
    parser.add_argument(
        "--exclude-email-fragment",
        action="append",
        default=[],
        help="Email substring to exclude; repeatable",
    )
    return parser.parse_args()


def load_config(path: Path | None) -> dict[str, Any]:
    if path is None:
        return {}
    if not path.exists():
        raise ConfigError(f"Config file not found: {path}")
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def workflow_settings(args: argparse.Namespace) -> dict[str, Any]:
    config = load_config(args.config)
    config_dir = args.config.parent if args.config else Path(".")
    presets = config.get("presets", {})
    preset = presets.get(args.preset, {})
    if args.preset and args.preset not in presets and args.config:
        raise ConfigError(f"Preset {args.preset!r} not found in {args.config}")

    defaults = config.get("defaults", {})
    rules = dict(DEFAULT_RULES)
    rules.update({key: value for key, value in defaults.items() if key in RULE_KEYS})
    rules.update({key: value for key, value in preset.items() if key in RULE_KEYS})
    for key in RULE_KEYS:
        value = getattr(args, key)
        if value is not None:
            rules[key] = value

    exclude_fragments = list(defaults.get("exclude_email_fragments", ["@silkresource"]))
    exclude_fragments.extend(args.exclude_email_fragment)

    columns = dict(DEFAULT_COLUMNS)
    columns.update(config.get("matrixify", {}).get("columns", {}))
    sheet_name = config.get("matrixify", {}).get("sheet_name", "Orders")

    files = config.get("files", {})
    run_dir = args.run
    input_path = args.input
    output_path = args.output
    csv_output_path = args.csv_output
    google_upload_output_path = args.google_upload_output
    pinterest_upload_output_path = args.pinterest_upload_output
    summary_output_path = args.summary_output

    if input_path is None and run_dir is not None:
        input_path = select_input_file(
            run_dir / files.get("source_dir", "source"),
            pattern=files.get("input_pattern", "Export_*.xlsx"),
        )

    if output_path is None and run_dir is not None:
        run_date = infer_run_date(run_dir, input_path)
        output_dir = run_dir / files.get("output_dir", "output")
        output_stem = output_filename_stem(
            prefix=files.get("output_prefix", "customer_match"),
            run_date=run_date,
            preset=args.preset,
        )
        output_path = output_dir / f"{output_stem}.xlsx"
        csv_output_path = csv_output_path or output_dir / f"{output_stem}.csv"
        google_upload_output_path = google_upload_output_path or output_dir / f"{output_stem}.google_upload.csv"
        pinterest_upload_output_path = pinterest_upload_output_path or output_dir / f"{output_stem}.pinterest_upload.csv"
        summary_output_path = summary_output_path or output_dir / f"{output_stem}.summary.json"

    if input_path is None or output_path is None:
        raise ConfigError("Provide --input and --output, or provide --run so they can be resolved from config.")

    return {
        "config_path": str(args.config) if args.config else None,
        "preset": args.preset,
        "preset_description": preset.get("description"),
        "rules": rules,
        "exclude_emails": {norm_email(email) for email in args.exclude_email},
        "exclude_fragments": tuple(fragment.lower() for fragment in exclude_fragments),
        "columns": columns,
        "sheet_name": sheet_name,
        "input": input_path,
        "output": output_path,
        "csv_output": csv_output_path,
        "google_upload_output": google_upload_output_path,
        "pinterest_upload_output": pinterest_upload_output_path,
        "summary_output": summary_output_path,
        "config_dir": config_dir,
    }


def select_input_file(source_dir: Path, *, pattern: str) -> Path:
    matches = sorted(source_dir.glob(pattern), key=lambda path: path.stat().st_mtime, reverse=True)
    if not matches:
        raise ConfigError(f"No input files matching {pattern!r} found in {source_dir}")
    return matches[0]


def infer_run_date(run_dir: Path, input_path: Path | None) -> str:
    if re.match(r"^\d{4}-\d{2}-\d{2}", run_dir.name):
        return run_dir.name[:10]
    if input_path:
        match = re.search(r"(\d{4}-\d{2}-\d{2})", input_path.name)
        if match:
            return match.group(1)
    return dt.date.today().isoformat()


def output_filename_stem(*, prefix: str, run_date: str, preset: str | None) -> str:
    if not preset or preset == "baseline":
        return f"{prefix}_{run_date}"
    return f"{prefix}_{run_date}_{preset}"


def require_columns(headers: list[Any], columns: dict[str, str]) -> dict[str, int]:
    idx = {header: position for position, header in enumerate(headers)}
    missing = [column for column in columns.values() if column not in idx]
    if missing:
        formatted = ", ".join(repr(column) for column in sorted(set(missing)))
        raise ConfigError(f"Input workbook is missing required columns: {formatted}")
    return idx


def main() -> None:
    args = parse_args()
    settings = workflow_settings(args)
    rules = settings["rules"]
    columns = settings["columns"]

    workbook = openpyxl.load_workbook(settings["input"], read_only=True, data_only=True)
    source = workbook[settings["sheet_name"]]
    headers = [source.cell(1, column).value for column in range(1, source.max_column + 1)]
    idx = require_columns(headers, columns)

    order_rows = []
    max_processed_at = None
    for row in source.iter_rows(min_row=2, values_only=True):
        if row_value(row, idx, columns, "top_row") not in TOP_VALUES:
            continue
        processed_at = parse_dt(row_value(row, idx, columns, "processed_at"))
        if processed_at and (max_processed_at is None or processed_at > max_processed_at):
            max_processed_at = processed_at
        order_rows.append(row)

    if max_processed_at is None:
        raise RuntimeError("No processed order dates found.")

    recent_cutoff = (max_processed_at - dt.timedelta(days=365 * rules["lookback_years"])).date()
    customers = defaultdict(
        lambda: {
            "orders": 0,
            "spend": 0.0,
            "max_order": 0.0,
            "first": None,
            "last": None,
            "tags": set(),
            "line_classes": Counter(),
            "lifetime_spent": 0.0,
            "lifetime_orders": 0,
        }
    )
    orders_by_id = {}
    valid_recent_orders = 0

    for row in order_rows:
        email = norm_email(row_value(row, idx, columns, "email") or row_value(row, idx, columns, "customer_email"))
        processed_at = parse_dt(row_value(row, idx, columns, "processed_at"))
        current_total = row_value(row, idx, columns, "current_total")
        total = money(current_total if current_total is not None else row_value(row, idx, columns, "total"))
        payment_status = str(row_value(row, idx, columns, "payment_status") or "").lower()

        if not email or not EMAIL_RE.match(email):
            continue
        if email in settings["exclude_emails"] or any(fragment in email for fragment in settings["exclude_fragments"]):
            continue
        if not processed_at or processed_at.date() < recent_cutoff:
            continue
        if row_value(row, idx, columns, "cancelled_at") or payment_status == "voided" or total <= 0:
            continue

        order_id = row_value(row, idx, columns, "order_id")
        if order_id in orders_by_id:
            continue

        valid_recent_orders += 1
        orders_by_id[order_id] = email
        customer = customers[email]
        customer["orders"] += 1
        customer["spend"] += total
        customer["max_order"] = max(customer["max_order"], total)
        customer["first"] = processed_at if customer["first"] is None or processed_at < customer["first"] else customer["first"]
        customer["last"] = processed_at if customer["last"] is None or processed_at > customer["last"] else customer["last"]
        customer["tags"].update(split_tags(row_value(row, idx, columns, "customer_tags")))
        customer["tags"].update(split_tags(row_value(row, idx, columns, "order_tags")))
        customer["lifetime_spent"] = max(customer["lifetime_spent"], money(row_value(row, idx, columns, "customer_total_spent")))
        customer["lifetime_orders"] = max(customer["lifetime_orders"], int(money(row_value(row, idx, columns, "customer_orders_count"))))

    for row in source.iter_rows(min_row=2, values_only=True):
        order_id = row_value(row, idx, columns, "order_id")
        if order_id not in orders_by_id or row_value(row, idx, columns, "line_type") != "Line Item":
            continue
        customers[orders_by_id[order_id]]["line_classes"].update(classify_line(row, idx, columns))

    selected = []
    for email, customer in customers.items():
        include = (
            customer["spend"] >= rules["min_recent_spend"]
            or customer["orders"] >= rules["repeat_order_min"]
            or customer["max_order"] >= rules["large_single_order"]
            or "designer-program" in customer["tags"]
            or (customer["lifetime_spent"] >= rules["lifetime_spend"] and customer["spend"] >= rules["lifetime_recent_spend"])
        )
        if include:
            selected.append(
                {
                    "email": email,
                    "reason": build_reason(customer, rules),
                    "spend": customer["spend"],
                    "orders": customer["orders"],
                    "last": customer["last"].date().isoformat(),
                }
            )

    selected.sort(key=lambda item: (-item["spend"], -item["orders"], item["email"]))
    total_selected_before_limit = len(selected)
    if rules["max_customers"] is not None:
        selected = selected[: int(rules["max_customers"])]

    output_workbook = openpyxl.Workbook()
    sheet = output_workbook.active
    sheet.title = "Customer Match"
    sheet.append(["email", "reason"])

    for item in selected:
        sheet.append([item["email"], item["reason"]])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:B{len(selected) + 1}"
    sheet["A1"].font = Font(bold=True)
    sheet["B1"].font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    sheet["A1"].fill = header_fill
    sheet["B1"].fill = header_fill
    sheet.column_dimensions[get_column_letter(1)].width = 34
    sheet.column_dimensions[get_column_letter(2)].width = 118

    settings["output"].parent.mkdir(parents=True, exist_ok=True)
    output_workbook.save(settings["output"])

    if settings["csv_output"]:
        settings["csv_output"].parent.mkdir(parents=True, exist_ok=True)
        with settings["csv_output"].open("w", newline="", encoding="utf-8") as file:
            writer = csv.DictWriter(file, fieldnames=["email", "reason"])
            writer.writeheader()
            for item in selected:
                writer.writerow({"email": item["email"], "reason": item["reason"]})

    if settings["google_upload_output"]:
        settings["google_upload_output"].parent.mkdir(parents=True, exist_ok=True)
        with settings["google_upload_output"].open("w", newline="", encoding="utf-8") as file:
            writer = csv.DictWriter(file, fieldnames=["Email"])
            writer.writeheader()
            for item in selected:
                writer.writerow({"Email": item["email"]})

    if settings["pinterest_upload_output"]:
        settings["pinterest_upload_output"].parent.mkdir(parents=True, exist_ok=True)
        with settings["pinterest_upload_output"].open("w", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            for item in selected:
                writer.writerow([item["email"]])

    summary = {
        "config": settings["config_path"],
        "preset": settings["preset"],
        "preset_description": settings["preset_description"],
        "input": str(settings["input"]),
        "output": str(settings["output"]),
        "csv_output": str(settings["csv_output"]) if settings["csv_output"] else None,
        "google_upload_output": str(settings["google_upload_output"]) if settings["google_upload_output"] else None,
        "pinterest_upload_output": str(settings["pinterest_upload_output"]) if settings["pinterest_upload_output"] else None,
        "latest_processed_order": max_processed_at.isoformat(sep=" "),
        "recent_cutoff": recent_cutoff.isoformat(),
        "top_order_rows": len(order_rows),
        "valid_recent_orders": valid_recent_orders,
        "valid_recent_customers": len(customers),
        "selected_customers": len(selected),
        "selected_customers_before_limit": total_selected_before_limit,
        "selected_recent_orders": sum(item["orders"] for item in selected),
        "selected_recent_spend": round(sum(item["spend"] for item in selected), 2),
        "rules": rules,
    }

    if settings["summary_output"]:
        settings["summary_output"].parent.mkdir(parents=True, exist_ok=True)
        settings["summary_output"].write_text(json.dumps(summary, indent=2), encoding="utf-8")

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
