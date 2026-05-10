#!/usr/bin/env python3
"""Build a two-column Google Customer Match review workbook from Shopify orders.

The source file is expected to be a Matrixify/Shopify order export with an
`Orders` sheet and standard order, customer, and line-item columns.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
TOP_VALUES = {True, "TRUE", "true", "Y", "Yes", 1, "1"}


def parse_dt(value: Any) -> dt.datetime | None:
    if not value:
        return None
    if isinstance(value, dt.datetime):
        return value
    return dt.datetime.strptime(str(value)[:19], "%Y-%m-%d %H:%M:%S")


def money(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def norm_email(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def split_tags(value: Any) -> set[str]:
    if not value:
        return set()
    return {part.strip().lower() for part in re.split(r"[,;]", str(value)) if part.strip()}


def classify_line(row: tuple[Any, ...], idx: dict[str, int]) -> set[str]:
    text = " ".join(
        str(row[idx[column]] or "").lower()
        for column in ("Line: Title", "Line: Name", "Line: SKU", "Line: Product Handle", "Line: Vendor")
    )
    line_total = money(row[idx["Line: Total"]])
    classes = set()

    if "sample" in text or (0 < line_total < 30):
        classes.add("sample")
    if "wallpaper" in text or "double roll" in text or "single roll" in text:
        classes.add("wallpaper")
    if any(term in text for term in ("fabric", "silk", "linen", "cotton", "velvet", "damask", "jacquard", "toile")):
        classes.add("fabric")
    if any(term in text for term in ("trim", "tape", "cord", "fringe", "samuel & sons")):
        classes.add("trim")
    if any(term in text for term in ("pillow", "bolster")):
        classes.add("pillow")
    if "scalamandre" in text or "old world weavers" in text:
        classes.add("scalamandre")

    return classes


def build_reason(customer: dict[str, Any]) -> str:
    reasons = []

    if "designer-program" in customer["tags"]:
        reasons.append("designer-program customer")
    if customer["spend"] >= 1000:
        reasons.append(f"high recent spend ${customer['spend']:,.0f}")
    elif customer["spend"] >= 250:
        reasons.append(f"qualified recent spend ${customer['spend']:,.0f}")
    elif customer["orders"] >= 2:
        reasons.append(f"repeat buyer with {customer['orders']} recent orders")
    elif customer["max_order"] >= 500:
        reasons.append(f"large single order ${customer['max_order']:,.0f}")

    if customer["lifetime_spent"] >= 3000 and customer["spend"] >= 100:
        reasons.append(f"strong lifetime value ${customer['lifetime_spent']:,.0f}")
    elif customer["lifetime_orders"] >= 3 and customer["spend"] >= 100:
        reasons.append(f"{customer['lifetime_orders']} lifetime orders")

    classes = customer["line_classes"]
    if classes["fabric"]:
        reasons.append("fabric buyer")
    if classes["wallpaper"]:
        reasons.append("wallpaper buyer")
    if classes["scalamandre"]:
        reasons.append("Scalamandre/design-house interest")
    if classes["trim"]:
        reasons.append("trim buyer")
    if classes["pillow"]:
        reasons.append("pillow buyer")

    reasons.append(f"last order {customer['last'].date().isoformat()}")
    return "; ".join(reasons)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path, help="Path to Matrixify/Shopify order export .xlsx")
    parser.add_argument("--output", required=True, type=Path, help="Path to write the two-column .xlsx output")
    parser.add_argument("--csv-output", type=Path, help="Optional two-column CSV output path")
    parser.add_argument("--summary-output", type=Path, help="Optional JSON summary output path")
    parser.add_argument("--lookback-years", type=int, default=3)
    parser.add_argument("--min-recent-spend", type=float, default=250)
    parser.add_argument("--repeat-order-min", type=int, default=2)
    parser.add_argument("--large-single-order", type=float, default=500)
    parser.add_argument("--lifetime-spend", type=float, default=3000)
    parser.add_argument("--lifetime-recent-spend", type=float, default=100)
    parser.add_argument("--exclude-email", action="append", default=[], help="Email address to exclude; repeatable")
    parser.add_argument(
        "--exclude-email-fragment",
        action="append",
        default=["@silkresource"],
        help="Email substring to exclude; repeatable",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    excluded_emails = {norm_email(email) for email in args.exclude_email}
    excluded_fragments = tuple(fragment.lower() for fragment in args.exclude_email_fragment)

    workbook = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    source = workbook["Orders"]
    headers = [source.cell(1, column).value for column in range(1, source.max_column + 1)]
    idx = {header: position for position, header in enumerate(headers)}

    order_rows = []
    max_processed_at = None
    for row in source.iter_rows(min_row=2, values_only=True):
        if row[idx["Top Row"]] not in TOP_VALUES:
            continue
        processed_at = parse_dt(row[idx["Processed At"]])
        if processed_at and (max_processed_at is None or processed_at > max_processed_at):
            max_processed_at = processed_at
        order_rows.append(row)

    if max_processed_at is None:
        raise RuntimeError("No processed order dates found.")

    recent_cutoff = max_processed_at - dt.timedelta(days=365 * args.lookback_years)
    customers = defaultdict(
        lambda: {
            "orders": 0,
            "spend": 0.0,
            "max_order": 0.0,
            "first": None,
            "last": None,
            "tags": set(),
            "line_classes": Counter(),
            "lifetime_spent": 0.0,
            "lifetime_orders": 0,
        }
    )
    orders_by_id = {}
    valid_recent_orders = 0

    for row in order_rows:
        email = norm_email(row[idx["Email"]] or row[idx["Customer: Email"]])
        processed_at = parse_dt(row[idx["Processed At"]])
        total = money(row[idx["Price: Current Total"]] if row[idx["Price: Current Total"]] is not None else row[idx["Price: Total"]])
        payment_status = str(row[idx["Payment: Status"]] or "").lower()

        if not email or not EMAIL_RE.match(email):
            continue
        if email in excluded_emails or any(fragment in email for fragment in excluded_fragments):
            continue
        if not processed_at or processed_at < recent_cutoff:
            continue
        if row[idx["Cancelled At"]] or payment_status == "voided" or total <= 0:
            continue

        order_id = row[idx["ID"]]
        if order_id in orders_by_id:
            continue

        valid_recent_orders += 1
        orders_by_id[order_id] = email
        customer = customers[email]
        customer["orders"] += 1
        customer["spend"] += total
        customer["max_order"] = max(customer["max_order"], total)
        customer["first"] = processed_at if customer["first"] is None or processed_at < customer["first"] else customer["first"]
        customer["last"] = processed_at if customer["last"] is None or processed_at > customer["last"] else customer["last"]
        customer["tags"].update(split_tags(row[idx["Customer: Tags"]]))
        customer["tags"].update(split_tags(row[idx["Tags"]]))
        customer["lifetime_spent"] = max(customer["lifetime_spent"], money(row[idx["Customer: Total Spent"]]))
        customer["lifetime_orders"] = max(customer["lifetime_orders"], int(money(row[idx["Customer: Orders Count"]])))

    for row in source.iter_rows(min_row=2, values_only=True):
        order_id = row[idx["ID"]]
        if order_id not in orders_by_id or row[idx["Line: Type"]] != "Line Item":
            continue
        customers[orders_by_id[order_id]]["line_classes"].update(classify_line(row, idx))

    selected = []
    for email, customer in customers.items():
        include = (
            customer["spend"] >= args.min_recent_spend
            or customer["orders"] >= args.repeat_order_min
            or customer["max_order"] >= args.large_single_order
            or "designer-program" in customer["tags"]
            or (customer["lifetime_spent"] >= args.lifetime_spend and customer["spend"] >= args.lifetime_recent_spend)
        )
        if include:
            selected.append(
                {
                    "email": email,
                    "reason": build_reason(customer),
                    "spend": customer["spend"],
                    "orders": customer["orders"],
                    "last": customer["last"].date().isoformat(),
                }
            )

    selected.sort(key=lambda item: (-item["spend"], -item["orders"], item["email"]))

    output_workbook = openpyxl.Workbook()
    sheet = output_workbook.active
    sheet.title = "Customer Match"
    sheet.append(["email", "reason"])

    for item in selected:
        sheet.append([item["email"], item["reason"]])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:B{len(selected) + 1}"
    sheet["A1"].font = Font(bold=True)
    sheet["B1"].font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    sheet["A1"].fill = header_fill
    sheet["B1"].fill = header_fill
    sheet.column_dimensions[get_column_letter(1)].width = 34
    sheet.column_dimensions[get_column_letter(2)].width = 118

    args.output.parent.mkdir(parents=True, exist_ok=True)
    output_workbook.save(args.output)

    if args.csv_output:
        args.csv_output.parent.mkdir(parents=True, exist_ok=True)
        with args.csv_output.open("w", newline="", encoding="utf-8") as file:
            writer = csv.DictWriter(file, fieldnames=["email", "reason"])
            writer.writeheader()
            for item in selected:
                writer.writerow({"email": item["email"], "reason": item["reason"]})

    summary = {
        "input": str(args.input),
        "output": str(args.output),
        "csv_output": str(args.csv_output) if args.csv_output else None,
        "latest_processed_order": max_processed_at.isoformat(sep=" "),
        "recent_cutoff": recent_cutoff.isoformat(sep=" "),
        "top_order_rows": len(order_rows),
        "valid_recent_orders": valid_recent_orders,
        "valid_recent_customers": len(customers),
        "selected_customers": len(selected),
        "selected_recent_orders": sum(item["orders"] for item in selected),
        "selected_recent_spend": round(sum(item["spend"] for item in selected), 2),
        "rules": {
            "lookback_years": args.lookback_years,
            "min_recent_spend": args.min_recent_spend,
            "repeat_order_min": args.repeat_order_min,
            "large_single_order": args.large_single_order,
            "lifetime_spend": args.lifetime_spend,
            "lifetime_recent_spend": args.lifetime_recent_spend,
        },
    }

    if args.summary_output:
        args.summary_output.parent.mkdir(parents=True, exist_ok=True)
        args.summary_output.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
